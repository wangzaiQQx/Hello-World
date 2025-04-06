import pandas as pd
from openpyxl import load_workbook
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

def prettify(element):
    """Return a pretty-printed XML string for the Element."""
    rough_string = tostring(element, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")

def get_cell_styles(ws, row, col):
    """
    获取单元格的样式信息。
    :param ws: 工作表对象
    :param row: 行号
    :param col: 列号
    :return: 字典，包含样式信息
    """
    cell = ws.cell(row=row, column=col)
    
    # 获取字体颜色
    font_color = cell.font.color.rgb if cell.font.color else None
    
    # 获取背景颜色
    background_color = cell.fill.start_color.rgb if cell.fill.start_color and cell.fill.start_color.rgb else None
    
    # 获取边框样式
    border_style = {
        'left': cell.border.left.style if cell.border and cell.border.left else None,
        'right': cell.border.right.style if cell.border and cell.border.right else None,
        'top': cell.border.top.style if cell.border and cell.border.top else None,
        'bottom': cell.border.bottom.style if cell.border and cell.border.bottom else None,
    }
    
    return {
        'font_color': font_color,
        'background_color': background_color,
        'border_style': border_style
    }

def excel_to_xml(excel_file, xml_file):
    # 加载Excel文件
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active  # 选择第一个工作表

    # 创建XML根节点
    root = Element('data')

    # 获取列名
    column_names = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 遍历Excel文件中的每一行
    for row_cells in ws.iter_rows(min_row=2, values_only=False):  # 使用values_only=False获取单元格对象
        record = SubElement(root, 'record')
        
        # 遍历每一列
        for col_index, cell in enumerate(row_cells, start=1):
            field = SubElement(record, 'field')
            field.set('name', column_names[col_index - 1])  # 设置列名
            field.text = str(cell.value) if cell.value is not None else ''
            
            # 获取单元格的样式信息
            styles = get_cell_styles(ws, cell.row, col_index)
            
            # 添加字体颜色
            if styles['font_color']:
                field.set('font_color', styles['font_color'])
            
            # 添加背景颜色
            if styles['background_color']:
                field.set('background_color', styles['background_color'])
            
            # 添加边框样式
            for side, style in styles['border_style'].items():
                if style:
                    field.set(f'border_{side}_style', style)

    # 将XML写入文件
    with open(xml_file, 'w', encoding='utf-8') as f:
        f.write(prettify(root))

# 使用方法
excel_file = 'example.xlsx'  # 替换为您的Excel文件路径
xml_file = 'output.xml'      # 替换为您想要保存的XML文件路径
excel_to_xml(excel_file, xml_file)